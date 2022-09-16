import os
import time
import requests
from selenium import webdriver
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from datetime import datetime
import pyautogui
import sys



# 크롬 드라이버 자동 업데이트

from webdriver_manager.chrome import ChromeDriverManager

pyautogui.FAILSAFE = False


# 상대경로 세팅
if getattr(sys, 'frozen', False):
    #test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(sys.executable))
else:
    #python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(__file__))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

err_log_path = program_directory + "/Job_Error_log.txt"
save_date = datetime.today().strftime("%Y%m%d_%H%M")
exl_name = program_directory + f"/job_{save_date}.xlsx"
exl_sample_name = program_directory + f"/job_crawling_sample.xlsx"

# 브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# 불필요한 에러 메시지 삭제
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.implicitly_wait(2) # 웹페이지가 로딩될때까지 2초 기다림
# driver.maximize_window() # 화면 최대화

# 크롤링 방지 설정을 undefined로 변경
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
    "source": """
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            })
            """
})


# 엑셀 세팅


wb = openpyxl.load_workbook(exl_sample_name)
ws = wb.active


# 메인 페이지 이동
wait = WebDriverWait(driver, 5)
url = "https://www.jobplanet.co.kr/job"
driver.get(url=url)
time.sleep(5)

# IT/인터넷 클릭
driver.find_element(By.CSS_SELECTOR, "#JobPostingApp > div.recruitment-navigation > div > div.item-nav__contents > ul > li:nth-child(2) > button").click()
time.sleep(5)

# 데이터 분석 클릭
driver.find_element(By.CSS_SELECTOR, "#JobPostingApp > div.recruitment-navigation > div > div.item-nav__contents > ul > li:nth-child(7) > button").click()
time.sleep(5)


# 스크롤 전 높이
before_h = driver.execute_script("return window.scrollY")

# 무한 스크롤
while True:
    # 맨 아래로 스크롤을 내린다.
    driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.END)

    # 스크롤 사이 페이지 로딩 시간
    time.sleep(1)

    # 스크롤 후 높이
    after_h = driver.execute_script("return window.scrollY")
    if after_h == before_h:
        break
    before_h = after_h





link_num = 1
exl_num = 1

# 채용정보 ids로 리스트화
ids = driver.find_elements(by = By.CSS_SELECTOR, value = ".item-card")
time.sleep(2)


# 채용정보별 데이터 수집
for id in ids :

    link = id.find_element(By.CSS_SELECTOR, f"#JobPostingApp > div.recruitment-content > div.recruitment-content__list > div.infinite-scroll-component__outerdiv > div > div > div:nth-child({link_num}) > a").get_attribute('href')
    target = id.find_element(By.CSS_SELECTOR, f"#JobPostingApp > div.recruitment-content > div.recruitment-content__list > div.infinite-scroll-component__outerdiv > div > div > div:nth-child({link_num}) > a")
    co_name = id.find_element(By.CSS_SELECTOR, f"#JobPostingApp > div.recruitment-content > div.recruitment-content__list > div.infinite-scroll-component__outerdiv > div > div > div:nth-child({link_num}) > a > div.item-card__information > div.item-card__company > div.item-card__name").text
    print(f"{link_num} : {co_name}")
    time.sleep(1)
    link_num += 1

    # 채용공고문 새탭으로 열기
    pyautogui.hotkey("command", "t")
    # driver.send_keys(Keys.command+"t")
    # target.send_Keys(Keys.COMMAND + "\t")
    time.sleep(1)  
    all_windows = driver.window_handles
    driver.switch_to.window(all_windows[1])
    driver.get(link)
    time.sleep(3)


    # 직군 수집
    try :
        jikgun = driver.find_element(By.CSS_SELECTOR, ".ttl").text
    except:
        jikgun = ""

    # 직무 수집
    try :    
        jikmu = driver.find_element(By.CSS_SELECTOR, "#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div.recruitment-detail__box.recruitment-summary > dl > dd:nth-child(4)").text
    except:
        jikmu = ""

    # 마감일자 수집
    try :
        closing_date = driver.find_element(By.CSS_SELECTOR, ".recruitment-summary__end").text
    except:
        closing_date = ""


    # 기타 데이터 찾기

    i = 1

    upmu = ""
    jakyuk = ""
    woodea = ""
    bokri = ""
    locate = ""


    while i < 13 :
        try :
            el_name = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div:nth-child({i}) > h3").text

            if el_name == "주요 업무" :

                upmu = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div:nth-child({i}) > p").text


            if el_name == "자격 요건" :

                jakyuk = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div:nth-child({i}) > p").text

 
            if el_name == "우대사항" :

                woodea = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div:nth-child({i}) > p").text

            if el_name == "복리후생" :

                bokri = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div:nth-child({i}) > p").text

            if el_name == "회사위치" :

                locate = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div:nth-child({i}) > p").text

            i += 1

        except :
            break


    # 담당자/연락처/이메일 찾기

    i = 2

    contact1 = ""
    contact2 = ""
    contact3 = ""

    while i < 7 :
        try :
            el_name = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div.recruitment-detail__box.recruitment-contact > dl:nth-child({i}) > dt").text


            if el_name == "담당자" :

                contact1 = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div.recruitment-detail__box.recruitment-contact > dl:nth-child({i}) > dd").text


            if el_name == "연락처" :

                contact2 = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div.recruitment-detail__box.recruitment-contact > dl:nth-child({i}) > dd").text

 
            if el_name == "이메일" :

                contact3 = driver.find_element(By.CSS_SELECTOR, f"#job_search_app > div > div.job_search_content > section > div.job_search_detail > div > div > div > div.job_wrap_new.company_job_details > div > div.wrap > div > div > div.block_job_posting > section > div.recruitment-detail__box.recruitment-contact > dl:nth-child({i}) > dd").text
 

            i += 1

        except :
            break

    contact = f"{contact1}\n{contact2}\n{contact3}"        

    # 엑셀 저장

    ws[f'A{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',jikgun))
    ws[f'B{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',co_name))
    ws[f'C{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',jikmu))
    ws[f'D{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',upmu))
    ws[f'E{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',jakyuk))
    ws[f'F{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',woodea))
    ws[f'G{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',bokri))
    ws[f'H{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',closing_date))
    ws[f'I{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',locate))
    ws[f'J{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',contact))
    ws[f'K{exl_num+1}'] = str(ILLEGAL_CHARACTERS_RE.sub(r'',link))


    wb.save(exl_name)
    print("저장완료")
    exl_num += 1
    time.sleep(1)


    # 드라이브 닫기
    driver.close()
    driver.switch_to.window(all_windows[0])


print(f"총 {exl_num-1}개의 데이터가 저장되었습니다.")


while(True):
    pass
